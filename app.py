# pyrefly: ignore [missing-import]
import streamlit as st
import asyncio
import io
import zipfile
import requests
import re
import json
import openpyxl
from scraper import scrape_amazon_images
import subprocess

# Ensure Playwright browsers are installed on startup (crucial for Streamlit Cloud deployment)
@st.cache_resource
def install_playwright_browsers():
    try:
        # Try running playwright install chromium directly
        subprocess.run(["playwright", "install", "chromium"], check=True)
    except Exception as e:
        # Fallback to running via python module if playwright is not in PATH
        try:
            subprocess.run(["python", "-m", "playwright", "install", "chromium"], check=True)
        except Exception as e2:
            st.error(f"Failed to install Playwright browsers: {e2}")

install_playwright_browsers()

# Page Config
st.set_page_config(
    page_title="Amazon High-Res Image Scraper",
    page_icon="📸",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom Styling (Glassmorphism & Neon Dark Theme)
st.markdown("""
<style>
    /* Dark Theme Core overrides */
    .stApp {
        background-color: #0d0e15;
        color: #e2e8f0;
    }
    
    /* Header styling */
    .header-container {
        text-align: center;
        padding: 2.5rem 1rem;
        background: linear-gradient(135deg, rgba(26, 27, 46, 0.7) 0%, rgba(15, 16, 28, 0.7) 100%);
        border-radius: 16px;
        border: 1px solid rgba(255, 255, 255, 0.08);
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        margin-bottom: 2rem;
    }
    
    .header-title {
        font-family: 'Outfit', 'Inter', sans-serif;
        font-size: 2.8rem;
        font-weight: 800;
        background: linear-gradient(90deg, #a855f7 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
    
    .header-subtitle {
        font-family: 'Inter', sans-serif;
        font-size: 1.1rem;
        color: #94a3b8;
    }
    
    /* Cards and input container */
    .search-card {
        background: rgba(20, 21, 38, 0.6);
        border: 1px solid rgba(255, 255, 255, 0.05);
        border-radius: 12px;
        padding: 2rem;
        box-shadow: 0 4px 24px 0 rgba(0, 0, 0, 0.2);
        margin-bottom: 2rem;
    }
    
    /* Image Grid & Card Style */
    .img-card {
        background: rgba(30, 31, 54, 0.4);
        border: 1px solid rgba(255, 255, 255, 0.06);
        border-radius: 10px;
        padding: 10px;
        text-align: center;
        transition: transform 0.2s, border-color 0.2s;
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        margin-bottom: 15px;
    }
    
    .img-card:hover {
        transform: translateY(-4px);
        border-color: rgba(168, 85, 247, 0.5);
        box-shadow: 0 6px 20px rgba(168, 85, 247, 0.15);
    }
    
    /* Buttons and controls */
    .stButton>button {
        background: linear-gradient(90deg, #a855f7 0%, #3b82f6 100%) !important;
        color: white !important;
        border: none !important;
        padding: 0.5rem 2rem !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.3s !important;
        box-shadow: 0 4px 15px rgba(168, 85, 247, 0.3) !important;
    }
    
    .stButton>button:hover {
        opacity: 0.95 !important;
        transform: scale(1.02) !important;
        box-shadow: 0 6px 20px rgba(168, 85, 247, 0.5) !important;
    }
    
    /* Colour Tag badges */
    .color-badge {
        display: inline-block;
        padding: 0.3rem 0.8rem;
        background: rgba(168, 85, 247, 0.15);
        border: 1px solid rgba(168, 85, 247, 0.3);
        color: #d8b4fe;
        border-radius: 20px;
        font-size: 0.9rem;
        font-weight: 600;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# App Header
st.markdown("""
<div class="header-container">
    <div class="header-title">Amazon High-Res Image Scraper</div>
    <div class="header-subtitle">Extract and download high-resolution images for all colour variants of any Amazon product</div>
</div>
""", unsafe_allow_html=True)

# Helper function to compile and zip images
@st.cache_data(show_spinner=False)
def download_and_zip_images(results: dict, source_url: str) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Save the scraped source URL inside the ZIP file for future reference
        zip_file.writestr("scraped_url.txt", source_url)
        
        # Save the structured image URLs JSON inside the ZIP file as well
        zip_file.writestr("scraped_images.json", json.dumps(results, indent=2, ensure_ascii=False))
        
        for colour, urls in results.items():
            # Clean colour name to prevent file path issues
            safe_colour = re.sub(r'[\\/*?:"<>|]', "", colour).strip()
            safe_colour = safe_colour.replace(" ", "_")
            if not safe_colour:
                safe_colour = "Variant"
                
            for idx, url in enumerate(urls, start=1):
                try:
                    response = requests.get(url, timeout=10)
                    if response.status_code == 200:
                        # Determine file extension
                        ext = "jpg"
                        if ".png" in url.lower():
                            ext = "png"
                        elif ".webp" in url.lower():
                            ext = "webp"
                        elif ".gif" in url.lower():
                            ext = "gif"
                            
                        filename = f"{safe_colour}_{idx}.{ext}"
                        zip_file.writestr(filename, response.content)
                except Exception:
                    pass
    return zip_buffer.getvalue()

# Initialize session state variables
if "scraped_data" not in st.session_state:
    st.session_state["scraped_data"] = None
if "scraped_url" not in st.session_state:
    st.session_state["scraped_url"] = ""
if "proxy_url" not in st.session_state:
    st.session_state["proxy_url"] = ""

# Input Container
st.markdown('<div class="search-card">', unsafe_allow_html=True)
url_input = st.text_input(
    "Amazon Product URL",
    placeholder="https://www.amazon.co.uk/dp/B0D8THGP6G?th=1",
    value=st.session_state["scraped_url"] if st.session_state["scraped_url"] else "",
    help="Paste the full Amazon product link here (handles amazon.com, amazon.co.uk, amazon.in, etc.)"
)

# Optional Proxy URL input
with st.expander("🌐 Proxy Settings (Highly Recommended for Streamlit Cloud)"):
    st.markdown("""
    💡 **Why do I need a proxy?** 
    Amazon actively blocks public cloud servers like Streamlit Cloud (AWS). To scrape directly from the cloud, you can route your requests through a residential proxy.
    """)
    proxy_input = st.text_input(
        "Residential Proxy URL",
        placeholder="http://username:password@proxy.example.com:8000",
        value=st.session_state["proxy_url"] if st.session_state["proxy_url"] else "",
        help="Supports HTTP/HTTPS/SOCKS5 formats: e.g., http://user:pass@host:port"
    )
    st.session_state["proxy_url"] = proxy_input
st.markdown('</div>', unsafe_allow_html=True)

col1, col2 = st.columns([1, 4])
with col1:
    scrape_clicked = st.button("Scrape Images", use_container_width=True)

# Scrape logic execution
if scrape_clicked:
    if not url_input:
        st.warning("Please enter a valid Amazon product URL first.")
        st.session_state["scraped_data"] = None
        st.session_state["scraped_url"] = ""
    else:
        st.session_state["scraped_url"] = url_input
        st.session_state["scraped_data"] = None # Clear old data
        
        # Async run wrapper
        async def run_scraper(url):
            proxy_url = st.session_state.get("proxy_url", "").strip()
            return await scrape_amazon_images(url, proxy_url=proxy_url if proxy_url else None)

        with st.spinner("🕷️ Starting browser session & loading Amazon... This may take up to a minute."):
            try:
                data = asyncio.run(run_scraper(url_input))
                st.session_state["scraped_data"] = data
                if data:
                    # Save a local copy in the workspace, exactly as the CLI did
                    with open("scraped_images.json", "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)
            except Exception as e:
                st.error(f"Failed to scrape: {e}")
                st.session_state["scraped_data"] = None

# Render results from session state if available
if st.session_state["scraped_data"] is not None:
    data = st.session_state["scraped_data"]
    url_input = st.session_state["scraped_url"]
    
    if not data:
        st.error("No image variants could be extracted. Check the URL or try again later.")
        st.info("💡 Tip: Amazon sometimes displays a CAPTCHA. If this happens consistently, the server's IP address might be blocked by Amazon.")
    else:
        st.write("---")
        total_images = sum(len(urls) for urls in data.values())
        st.success(f"Successfully scraped {len(data)} colour variant(s) with {total_images} total image(s)!")
        
        # Zip all images helper
        with st.spinner("📦 Fetching high-res image files to build your ZIP archive..."):
            zip_data = download_and_zip_images(data, url_input)
            
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📥 Download All Images (.ZIP)",
                data=zip_data,
                file_name="amazon_scraped_images.zip",
                mime="application/zip",
                use_container_width=True
            )
        with col_dl2:
            st.download_button(
                label="📥 Download JSON Results (.JSON)",
                data=json.dumps(data, indent=2, ensure_ascii=False),
                file_name="scraped_images.json",
                mime="application/json",
                use_container_width=True
            )
        
        st.write("")
        
        # Display color categories
        for colour, urls in data.items():
            st.markdown(f'<div class="color-badge">{colour} ({len(urls)} images)</div>', unsafe_allow_html=True)
            
            # Image Grid (using 4 columns)
            cols = st.columns(4)
            for idx, url in enumerate(urls):
                col_idx = idx % 4
                with cols[col_idx]:
                    # Wrap in custom CSS card
                    st.markdown(f"""
                    <div class="img-card">
                        <img src="{url}" style="width:100%; border-radius:6px; max-height:220px; object-fit:contain; margin-bottom:10px;">
                        <div style="font-size:0.8rem; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; color:#94a3b8; margin-bottom:8px;">Image #{idx+1}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Direct Actions
                    c1, c2 = st.columns(2)
                    with c1:
                        st.link_button("🔗 Link", url, use_container_width=True)
                    with c2:
                        # Let user download individual file
                        try:
                            # Standard name
                            ext = "jpg"
                            if ".png" in url.lower(): ext = "png"
                            elif ".webp" in url.lower(): ext = "webp"
                            
                            # Fetch image content for single download
                            response = requests.get(url, timeout=5)
                            if response.status_code == 200:
                                st.download_button(
                                    label="💾 Save",
                                    data=response.content,
                                    file_name=f"{colour}_{idx+1}.{ext}",
                                    mime=f"image/{ext}",
                                    use_container_width=True,
                                    key=f"dl_{re.sub(r'[^a-zA-Z0-9]', '', colour)}_{idx}"
                                )
                            else:
                                st.button("❌", disabled=True, use_container_width=True, key=f"err_btn_{re.sub(r'[^a-zA-Z0-9]', '', colour)}_{idx}")
                        except Exception:
                            st.button("❌", disabled=True, use_container_width=True, key=f"exc_btn_{re.sub(r'[^a-zA-Z0-9]', '', colour)}_{idx}")
            
            st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
# EXCEL FILLER SECTION
# ─────────────────────────────────────────────────────────────────

def fill_excel_preserve_formatting(original_bytes, cell_updates):
    """
    Modify specific cells in the Template sheet of an xlsx file using openpyxl.
    This is simple and standard, but note that advanced Excel features like 
    dropdowns/data validations may be removed by openpyxl when saving.
    """
    input_buffer = io.BytesIO(original_bytes)
    wb = openpyxl.load_workbook(input_buffer)
    
    # Find the Template sheet
    template_sheet = None
    for name in wb.sheetnames:
        if "template" in name.lower():
            template_sheet = wb[name]
            break
            
    if template_sheet is None:
        if len(wb.sheetnames) > 0:
            template_sheet = wb[wb.sheetnames[0]]
        else:
            raise ValueError("No sheets found in workbook")
        
    # Apply cell updates
    for (row, col), value in cell_updates.items():
        template_sheet.cell(row=row, column=col, value=value)
        
    # Save to buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    wb.close()
    return output_buffer.getvalue()


st.write("")
st.write("---")
st.markdown("""
<div class="header-container" style="padding: 1.5rem 1rem;">
    <div class="header-title" style="font-size: 2rem;">📊 Excel Image Filler</div>
    <div class="header-subtitle">Upload your Temu Excel template and fill image URLs automatically by matching colour names</div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="search-card">', unsafe_allow_html=True)

# Step 1: Upload Excel
uploaded_excel = st.file_uploader(
    "Upload your Temu Excel template (.xlsx)",
    type=["xlsx"],
    help="Upload the Holdbacks/product Excel file. The app will match colour names and fill in image URLs."
)

# Step 2: Upload or use existing JSON
json_source = st.radio(
    "Image URL source",
    ["Use last scraped results (scraped_images.json)", "Upload a JSON file"],
    horizontal=True
)

uploaded_json = None
if json_source == "Upload a JSON file":
    uploaded_json = st.file_uploader("Upload scraped_images.json", type=["json"], key="json_uploader")

st.markdown('</div>', unsafe_allow_html=True)

# Step 3: Configuration
COLOUR_COL = 83   # Column CE — colour name
IMG_START_COL = 96  # Column CR — first image URL slot
IMG_END_COL = 105   # Column DA — last image URL slot (10 slots)
DATA_START_ROW = 5  # First data row in the Template sheet

fill_clicked = st.button("🔄 Fill Image URLs into Excel", use_container_width=False)

if fill_clicked:
    # Load JSON data
    scraped_data = None
    
    if json_source == "Upload a JSON file":
        if uploaded_json is None:
            st.warning("Please upload a JSON file first.")
            st.stop()
        try:
            scraped_data = json.loads(uploaded_json.read().decode("utf-8"))
        except Exception as e:
            st.error(f"Failed to parse JSON file: {e}")
            st.stop()
    else:
        # Try to load from local scraped_images.json
        try:
            with open("scraped_images.json", "r", encoding="utf-8") as f:
                scraped_data = json.load(f)
        except FileNotFoundError:
            st.error("No `scraped_images.json` file found. Please scrape a product first using the Single URL Scraper above, or upload a JSON file.")
            st.stop()
        except Exception as e:
            st.error(f"Failed to read scraped_images.json: {e}")
            st.stop()
    
    if not scraped_data:
        st.error("The JSON data is empty. Please scrape a product first.")
        st.stop()
    
    if uploaded_excel is None:
        st.warning("Please upload an Excel file first.")
        st.stop()
    
    # Read the original file bytes (we'll need them for the ZIP approach)
    original_bytes = uploaded_excel.getvalue()
    
    # Use openpyxl READ-ONLY to match colours and determine which cells to update
    try:
        wb = openpyxl.load_workbook(io.BytesIO(original_bytes), read_only=True)
    except Exception as e:
        st.error(f"Failed to open Excel file: {e}")
        st.stop()
    
    # Find the Template sheet
    template_sheet = None
    for name in wb.sheetnames:
        if "template" in name.lower():
            template_sheet = wb[name]
            break
    
    if template_sheet is None:
        if len(wb.sheetnames) > 0:
            st.info(f"ℹ️ Could not find a sheet named 'Template'. Falling back to the first sheet: **{wb.sheetnames[0]}**")
            template_sheet = wb[wb.sheetnames[0]]
        else:
            st.error("No sheets found in workbook")
            st.stop()
    
    # Build a lookup: normalised colour name -> list of clean image URLs
    def normalise(name):
        """Lowercase, strip spaces."""
        return re.sub(r'\s+', ' ', name.strip().lower())
    
    scraped_lookup = {}
    for colour_key, urls in scraped_data.items():
        norm_key = normalise(colour_key)
        from scraper import clean_image_url
        cleaned = [clean_image_url(u) for u in urls if u]
        scraped_lookup[norm_key] = cleaned
    
    # Determine which cells to update
    cell_updates = {}
    matched_count = 0
    unmatched_colours = []
    filled_rows = []
    
    for row in template_sheet.iter_rows(min_row=DATA_START_ROW, max_col=COLOUR_COL, values_only=False):
        cell = row[COLOUR_COL - 1]  # 0-indexed in iter_rows
        if cell.value is None:
            continue
        
        excel_colour = str(cell.value).strip()
        norm_excel = normalise(excel_colour)
        row_num = cell.row
        
        matched_urls = scraped_lookup.get(norm_excel)
        
        if matched_urls:
            for i, url in enumerate(matched_urls):
                col = IMG_START_COL + i
                if col > IMG_END_COL:
                    break
                cell_updates[(row_num, col)] = url
            matched_count += 1
            filled_rows.append((row_num, excel_colour, min(len(matched_urls), IMG_END_COL - IMG_START_COL + 1)))
        else:
            unmatched_colours.append((row_num, excel_colour))
    
    wb.close()
    
    # Show results
    if matched_count > 0:
        st.success(f"✅ Successfully matched images for **{matched_count}** colour variant(s)!")
        
        with st.expander("View matched rows", expanded=True):
            for row, colour, count in filled_rows:
                st.write(f"Row {row}: **{colour}** → {count} image(s) filled")
    
    if unmatched_colours:
        st.warning(f"⚠️ {len(unmatched_colours)} colour(s) in Excel had **no match** in the scraped data:")
        with st.expander("View unmatched colours"):
            for row, colour in unmatched_colours:
                st.write(f"Row {row}: **{colour}**")
        
        with st.expander("Available colours in scraped data"):
            for key in scraped_data.keys():
                st.write(f"• {key}")
    
    if matched_count == 0 and unmatched_colours:
        st.error("No colours could be matched. Make sure the colour names in your Excel file match the scraped colour names exactly.")
    
    # Generate the updated Excel
    if cell_updates:
        with st.spinner("📝 Writing image URLs into Excel..."):
            try:
                updated_bytes = fill_excel_preserve_formatting(original_bytes, cell_updates)
            except Exception as e:
                st.error(f"Failed to write Excel file: {e}")
                st.stop()
        
        st.download_button(
            label="📥 Download Updated Excel File",
            data=updated_bytes,
            file_name="Holdbacks_TEMU_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False
        )

